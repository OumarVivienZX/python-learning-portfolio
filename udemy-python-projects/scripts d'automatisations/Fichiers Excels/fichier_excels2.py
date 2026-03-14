import openpyxl

wb1 = openpyxl.load_workbook("octobre.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("novembre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("decembre.xlsx", data_only=True)



def ajouter_data_depuis_wb(wb,donnees):

    sheet = wb.active
    for row in range (2, sheet.max_row):
        liste_articles = sheet.cell(row, 1).value
        if not liste_articles:
            break
        liste_ventes = sheet.cell(row, 4).value
        if donnees.get(liste_articles):# get sert à récupérer la valeur associée à cette clé. Si la clé existe, get renvoie sa valeur, sinon elle renvoie None ou une valeur par défaut spécifiée.
            donnees[liste_articles].append(liste_ventes)
        else:
             donnees[liste_articles] = [liste_ventes]


data = {}

ajouter_data_depuis_wb(wb1, data)
ajouter_data_depuis_wb(wb2, data)
ajouter_data_depuis_wb(wb3, data)
print(data)

wb_sortie = openpyxl.Workbook()
sheet = wb_sortie.active

sheet["A1"] = "Article"
sheet["B1"] = "Octobre"
sheet["C1"] = "Novembre"
sheet["D1"] = "Décembre"

row = 2
for i in data.items() :
    nom_article = i[0]# i[0] est l'article, c'est la clé du dictionnaire data
    ventes = i[1]# i[1] est la liste des ventes associée à l'article i[0]
    #sheet.append([liste_article] + liste_ventes) # la méthode append ajoute une ligne à la feuille de calcul. En utilisant [liste_article] + liste_ventes, on crée une liste qui contient l'article suivi des ventes associées, ce qui correspond à la structure des colonnes définies précédemment.
    sheet.cell(row, 1).value = nom_article
    for j in range(0,len(ventes)):
        sheet.cell(row, j+2).value = ventes[j]# j+2 est utilisé pour placer les ventes dans les colonnes B, C, D, etc., en fonction de l'index j qui commence à 0. Ainsi, la première vente (j=0) sera placée en colonne B (j+2=2), la deuxième vente (j=1) en colonne C (j+2=3), et ainsi de suite.
        # vente[j] est la vente associée à l'article pour le mois correspondant à l'index j. En utilisant sheet.cell(row, j+2).value, on attribue cette vente à la cellule appropriée dans la ligne actuelle (row) et la colonne correspondante (j+2).
    row += 1
# ajout de graphique pour le total des ventes de chaque article
chart = openpyxl.chart.BarChart()# BarChart() crée un objet de type graphique à barres, qui est utilisé pour visualiser les données de manière graphique.
chart.title = "Total des ventes par article"
chart.x_axis.title = "Articles"# x_axis.title définit le titre de l'axe des x du graphique, qui dans ce cas est "Articles".chart.y_axis.title = "Total des ventes" 
data_chart = openpyxl.chart.Reference(sheet, min_col=2, max_col=4, min_row=1, max_row=row-1)
categories = openpyxl.chart.Reference(sheet, min_col=1, min_row=2, max_row=row-1)
chart.add_data(data_chart, titles_from_data=True)   
chart.set_categories(categories)
sheet.add_chart(chart, "F2")# F2 est la position où le graphique sera placé dans la feuille de calcul.and

    
wb_sortie.save("Fichier_sortie.xlsx")