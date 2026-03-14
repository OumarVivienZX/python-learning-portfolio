import openpyxl

wb = openpyxl.load_workbook("octobre.xlsx")

print(wb.sheetnames)# avoir le nombre de feuilles sur le livre 

sheet = wb.active # pour accéder à la feuille active du classeur une feuille est active par défaut c'est la première feuille du classeur sheet = feuille1

# cell = sheet["A1"] # pour accéder à la cellule A1
# print(cell.value) # pour afficher la valeur de la cellule A1
# cell = sheet["A2"] # pour accéder à la cellule A2
# print(cell.value) # pour afficher la valeur de la cellule A2

print (sheet.max_row) # pour afficher le nombre de lignes utilisées dans la feuille
print (sheet.max_column) # pour afficher le nombre de colonnes utilisées dans la feuille

for row in range (2, sheet.max_row ): # pour parcourir les lignes de la feuille à partir de la ligne 2 jusqu'à la dernière ligne utilisée
    for col in range (1, sheet.max_column + 1): # pour parcourir les colonnes de la feuille à partir de la colonne 1 jusqu'à la dernière colonne utilisée
        cell = sheet.cell(row=row, column=col) # pour accéder à la cellule à l'intersection de la ligne et de la colonne
        print(cell.value) # pour afficher la valeur de la cellule
    
